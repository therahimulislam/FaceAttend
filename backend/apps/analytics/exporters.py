"""
FaceAttend — Phase 15: Report Export Helpers

Provides three exporters:
  - csv_response(rows, headers, filename) → StreamingHttpResponse (text/csv)
  - excel_response(rows, headers, filename, sheet_name) → HttpResponse (.xlsx)
  - pdf_response(title, headers, rows, meta_lines, filename) → HttpResponse (.pdf)
"""
import csv
import io
from datetime import date

from django.http import HttpResponse, StreamingHttpResponse


# ---------------------------------------------------------------------------
# Brand colours (used in PDF / Excel headers)
# ---------------------------------------------------------------------------
BRAND_DARK = (15, 23, 42)      # slate-900
BRAND_MID = (30, 41, 59)       # slate-800
BRAND_ACCENT = (99, 102, 241)  # indigo-500
BRAND_TEXT = (248, 250, 252)   # slate-50


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

class _EchoBuffer:
    """Pseudo-buffer that echoes values for StreamingHttpResponse."""
    def write(self, value):
        return value


def csv_response(rows: list[list], headers: list[str], filename: str) -> StreamingHttpResponse:
    """
    Return a streaming CSV response.
    rows: list of lists (each inner list = one row of data).
    """
    pseudo_buffer = _EchoBuffer()
    writer = csv.writer(pseudo_buffer)

    def generate():
        yield writer.writerow(headers)
        for row in rows:
            yield writer.writerow(row)

    response = StreamingHttpResponse(generate(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def excel_response(
    rows: list[list],
    headers: list[str],
    filename: str,
    sheet_name: str = "Report",
) -> HttpResponse:
    """
    Return an Excel (.xlsx) response using openpyxl.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row styling
    header_fill = PatternFill(
        start_color="0F1729",  # brand dark
        end_color="0F1729",
        fill_type="solid",
    )
    accent_fill = PatternFill(
        start_color="6366F1",  # indigo
        end_color="6366F1",
        fill_type="solid",
    )
    header_font = Font(bold=True, color="F8FAFC", size=10)
    thin_border = Border(
        left=Side(style="thin", color="1E293B"),
        right=Side(style="thin", color="1E293B"),
        top=Side(style="thin", color="1E293B"),
        bottom=Side(style="thin", color="1E293B"),
    )

    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = accent_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 20

    # Data rows
    alt_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    normal_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    data_font = Font(color="E2E8F0", size=10)

    for i, row in enumerate(rows, start=2):
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else normal_fill
        for cell in ws[i]:
            cell.fill = fill
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-width columns
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(headers[col_idx - 1])),
            *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
            10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def pdf_response(
    title: str,
    headers: list[str],
    rows: list[list],
    meta_lines: list[str] | None = None,
    filename: str = "report.pdf",
) -> HttpResponse:
    """
    Return a PDF response using ReportLab.
    meta_lines: optional list of strings shown below the title (e.g. period, student name).
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = io.BytesIO()
    use_landscape = len(headers) > 7

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4) if use_landscape else A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    brand_indigo = colors.HexColor("#6366F1")
    brand_dark = colors.HexColor("#0F172A")
    brand_mid = colors.HexColor("#1E293B")
    brand_text = colors.HexColor("#E2E8F0")

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=brand_indigo,
        spaceAfter=6,
        alignment=TA_LEFT,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#94A3B8"),
        spaceAfter=3,
    )

    story = []

    # Title
    story.append(Paragraph(f"FaceAttend — {title}", title_style))
    story.append(Paragraph(f"Generated: {date.today().strftime('%d %B %Y')}", meta_style))
    if meta_lines:
        for line in meta_lines:
            story.append(Paragraph(line, meta_style))
    story.append(Spacer(1, 0.4 * cm))

    # Table
    page_width = (landscape(A4)[0] if use_landscape else A4[0]) - 3 * cm
    col_width = page_width / len(headers)

    table_data = [headers] + rows
    table = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), brand_indigo),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [brand_dark, brand_mid]),
        ("TEXTCOLOR", (0, 1), (-1, -1), brand_text),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#334155")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
